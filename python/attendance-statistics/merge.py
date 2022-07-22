#! /usr/bin/env python
# -*- coding: utf-8 -*-


import logging
import os
import openpyxl
import sys

from constant import *

# pathname filename
logging.basicConfig(format='%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s',
                    level=logging.INFO)
logger = logging.getLogger(__name__)

if __name__ == '__main__':
    file_handwork = os.path.join(os.getcwd(), FILE_DIR, FILE_HANDWORK)
    logger.info("[%s] file path: %s", FILE_HANDWORK, file_handwork)
    wb_handwork = openpyxl.load_workbook(file_handwork)
    sheet_handwork = wb_handwork.active
    row_num_handwork = sheet_handwork.max_row
    column_num_handwork = sheet_handwork.max_column
    logger.info("[%s] row number: %d, column number: %d", FILE_HANDWORK, row_num_handwork, column_num_handwork)
    valid_column_handwork = column_num_handwork - HANDWORK_START_COLUMN_INDEX - 3
    logger.info("[%s] valid day: %d", FILE_HANDWORK, valid_column_handwork)

    file_program = os.path.join(os.getcwd(), FILE_DIR, FILE_PROGRAM)
    logger.info("[%s] file path: %s", FILE_PROGRAM, file_program)
    wb_program = openpyxl.load_workbook(file_program)
    sheet_program = wb_program.active
    row_num_program = sheet_program.max_row
    column_num_program = sheet_program.max_column
    logger.info("[%s] row number: %d, column number: %d", FILE_PROGRAM, row_num_program, column_num_program)
    valid_column_program = 0
    rows_first = list(sheet_program.rows)[0]
    for index in range(0, len(rows_first)):
        column_value = rows_first[index].value
        # logger.info("%s", column_value)
        if column_value == CONTENT_WEEKDAY:
            logger.info("[%s] last valid data index: %d", FILE_PROGRAM, index - 1)
            valid_column_program = index - PROGRAM_START_COLUMN_INDEX
            break
    logger.info("[%s] valid day: %d", FILE_PROGRAM, valid_column_program)

    if valid_column_program != valid_column_handwork:
        logger.warning("valid day number is not equal: %s -- %d, %s -- %d", FILE_HANDWORK, valid_column_program,
                       FILE_PROGRAM, valid_column_handwork)
        sys.exit(1)

    # 字典:姓名-行号
    name_row_dict = {}
    handwork_rows_list = list(sheet_handwork.rows)
    index = 0
    for row in handwork_rows_list[1:]:
        index = index + 1
        name_row_dict[row[HANDWORK_NAME_COLUMN_INDEX].value] = index
    for name, index in name_row_dict.items():
        logger.debug("name: {}, row num: {}".format(name, index))

    program_rows_list = list(sheet_program.rows)[1:]
    for program_row in program_rows_list:
        name = program_row[PROGRAM_NAME_COLUMN_INDEX].value
        if name not in name_row_dict:
            logger.warning("%s not found in %s", name, FILE_HANDWORK)
            continue
        logger.debug("start to merge %s", name)
        for index in range(valid_column_program):
            time_value = program_row[PROGRAM_START_COLUMN_INDEX + index].value
            logger.debug("column: %d, time_value: %s", PROGRAM_START_COLUMN_INDEX + index, time_value)
            if isinstance(time_value, str):
                time_value = float(time_value)
            if time_value is not None and time_value > 0:
                row_num = name_row_dict[name]
                handwork_row = handwork_rows_list[row_num]
                handwork_row[HANDWORK_START_COLUMN_INDEX + index].value = time_value

    file_merged = os.path.join(os.getcwd(), FILE_DIR, FILE_MERGED)
    wb_handwork.save(file_merged)
    logger.info("合并结果写入文件: %s", file_merged)
