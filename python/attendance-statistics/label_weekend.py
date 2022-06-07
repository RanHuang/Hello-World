#! /usr/bin/env python
# -*- coding: utf-8 -*-


import logging
import os
import openpyxl

from constant import *

# pathname filename
logging.basicConfig(format='%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s',
                    level=logging.INFO)
logger = logging.getLogger(__name__)


if __name__ == '__main__':
    file_excel_origin = os.path.join(os.getcwd(), FILE_DIR, FILE_ORIGIN)
    logger.info("file path: %s", file_excel_origin)
    wb = openpyxl.load_workbook(file_excel_origin)
    sheet = wb.active
    row_num = sheet.max_row
    column_num = sheet.max_column
    logger.info("row number: %d, column number: %d", row_num, column_num)

    # 有效数据列需排除第一列
    rows_list = list(sheet.rows)[1:]

    logger.debug("print data of the first row:")
    for cell in rows_list[0]:
        logger.debug("type: %s, data: %s", type(cell.value), cell.value)

    # 判断是否为周末
    for row in rows_list:
        check_in_date = row[ORIGIN_COLUMN_CHECK_IN_DATE].value
        day = check_in_date.date()
        # Monday == 0 ... Sunday == 6
        if day.weekday() in [5, 6]:
            logger.debug("check in weekend: %s", check_in_date)
            row[ORIGIN_COLUMN_IS_WEEKEND].value = STRING_YES

    file_calc_origin = os.path.join(os.getcwd(), FILE_DIR, FILE_ORIGIN_WEEKEND)
    wb.save(file_calc_origin)
    logger.info("第0次程序处理结果写入文件: %s", file_calc_origin)
