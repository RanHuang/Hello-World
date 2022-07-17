#! /usr/bin/env python
# -*- coding: utf-8 -*-

import logging
import os
import openpyxl

from constant import *

# pathname filename
logging.basicConfig(format='%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s',
                    level=logging.INFO)
logger = logging.getLogger(__name__)


if __name__ == '__main__':
    file_excel_origin = os.path.join(os.getcwd(), FILE_DIR, FILE_ORIGIN_LATE)
    logger.info("file path: %s", file_excel_origin)
    wb = openpyxl.load_workbook(file_excel_origin)
    sheet = wb.active
    row_num = sheet.max_row
    column_num = sheet.max_column
    logger.info("row number: %d, column number: %d", row_num, column_num)

    # 有效数据列需排除第一列
    rows_list = list(sheet.rows)[1:]

    logger.debug("print data of the first row:")
    for cell in rows_list[0]:
        logger.debug("type: %s, data: %s", type(cell.value), cell.value)

    # 提取加班打卡时间提取
    for row in rows_list:
        # 初始化备注列为空字符串
        row[ORIGIN_COLUMN_IS_BE_LATE].value = ""
        row[ORIGIN_COLUMN_LEAVE_EARLY].value = ""

        check_in_time = row[ORIGIN_COLUMN_CHECK_IN_TIME].value
        time_list = [str_time.strip() for str_time in check_in_time.split(TIME_SEPARATOR)]
        time_list = [time_str for time_str in time_list if len(time_str) > 0]
        logger.debug("check in time: %s", time_list)
        if len(time_list) > 0:
            # 凌晨加班判定
            if time_list[0] >= TIME_WEEKDAY_BE_LATE_TIME:
                row[ORIGIN_COLUMN_IS_BE_LATE].value = STRING_BE_LATE
            if time_list[-1] < TIME_WEEKDAY_LEAVE_EARLY_TIME:
                row[ORIGIN_COLUMN_LEAVE_EARLY].value = STRING_LEAVE_EARLY

    file_calc_late_early = os.path.join(os.getcwd(), FILE_DIR, FILE_LATE_OR_EARLY)
    wb.save(file_calc_late_early)
    logger.info("迟到早退程序处理结果写入文件: %s", file_calc_late_early)
