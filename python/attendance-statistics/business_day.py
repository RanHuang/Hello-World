#! /usr/bin/env python
# -*- coding: utf-8 -*-

import logging
import os
import openpyxl

from constant import *

# pathname filename
logging.basicConfig(format='%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s',
                    level=logging.INFO)
logger = logging.getLogger(__name__)

if __name__ == '__main__':
    file_excel_handwork = os.path.join(os.getcwd(), FILE_DIR, FILE_HANDWORK)
    logger.info("file path: %s", file_excel_handwork)
    wb = openpyxl.load_workbook(file_excel_handwork)
    sheet = wb.active
    row_num = sheet.max_row
    column_num = sheet.max_column
    logger.info("row number: %d, column number: %d", row_num, column_num)

    row_first = list(sheet.rows)[0]
    logger.debug("print data of the first row:")
    for cell in row_first:
        logger.debug("type: %s, data: %s", type(cell.value), cell.value)
    last_column_num = len(row_first) - 1
    logger.info("last column index: %d, value: %s", last_column_num, row_first[last_column_num].value)
    # 有效数据列需排除第一列
    rows_list = list(sheet.rows)[1:]

    for row in rows_list:
        yellow_num = 0
        for cell in row[8:]:
            color = cell.fill.fgColor.rgb
            # business day - cell has color
            if color == COLOR_YELLOW:
                yellow_num += 1
        logger.info("business day: %d", yellow_num)
        row[-1].value = yellow_num

    file_calc_business = os.path.join(os.getcwd(), FILE_DIR, FILE_ON_BUSINESS)
    wb.save(file_calc_business)
    logger.info("加班天数统计结果写入文件: %s", file_calc_business)
