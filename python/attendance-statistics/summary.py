#! /usr/bin/env python
# -*- coding: utf-8 -*-


import logging
import os
import openpyxl
import math

from constant import *

# pathname filename
logging.basicConfig(format='%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s',
                    level=logging.INFO)
logger = logging.getLogger(__name__)

if __name__ == '__main__':
    file_calc = os.path.join(os.getcwd(), FILE_DIR, FILE_CALC)
    logger.info("file path: %s", file_calc)
    wb = openpyxl.load_workbook(file_calc)
    sheet = wb.active
    row_num = sheet.max_row
    column_num = sheet.max_column
    logger.info("row number: %d, column number: %d", row_num, column_num)

    # 有效数据列需排除第一列
    rows_list = list(sheet.rows)[1:]

    # 将姓名、日期和加班时长转换为嵌套字典
    summary_dic = {}
    for row in rows_list:
        name = row[ORIGIN_COLUMN_NAME].value
        check_in_date = row[ORIGIN_COLUMN_CHECK_IN_DATE].value
        duration_value = row[ORIGIN_COLUMN_OVER_TIME_STATISTIC].value
        date_time_dic = summary_dic.setdefault(name, {})
        date_time_dic[check_in_date] = duration_value
    logger.debug("summary dic: %s", summary_dic)

    date_set = set()
    for row in rows_list:
        check_in_date = row[ORIGIN_COLUMN_CHECK_IN_DATE].value
        date_set.add(check_in_date)
    date_list = sorted(list(date_set))
    logger.debug("all check in date: %s", date_list)

    # 读取姓名列表
    file_summary = os.path.join(os.getcwd(), FILE_DIR, FILE_NAME_LIST)
    logger.info("file path: %s", file_summary)
    wb_summary = openpyxl.load_workbook(file_summary)
    sheet_summary = wb_summary.active
    row_num_summary = sheet_summary.max_row
    column_num_summary = sheet_summary.max_column
    logger.info("row number summary: %d, column number summary: %d", row_num_summary, column_num_summary)

    # 填充第一行：日期
    rows_first = list(sheet_summary.rows)[0]
    logger.debug("column number: %d", len(rows_first))
    for index in range(0, len(date_list)):
        date_index = index + 4
        rows_first[date_index].value = date_list[index].strftime('%m-%d')
    rows_first[len(date_list) + 4].value = CONTENT_WEEKDAY
    rows_first[len(date_list) + 4 + 1].value = CONTENT_WEEKEND
    # 有效数据列需排除第一列
    rows_name_list = list(sheet_summary.rows)[1:]
    for row in rows_name_list:
        name = row[SUMMARY_COLUMN_NAME].value
        logger.info("name: %s", name)
        date_time_dic = summary_dic.setdefault(name, {})
        for index in range(0, len(date_list)):
            date_index = index + 4
            row[date_index].value = date_time_dic.setdefault(date_list[index], 0)
    # 统计平时加班时长
    for row in rows_name_list:
        total_over_time = 0
        for index in range(0, len(date_list)):
            day = date_list[index].date()
            if day.weekday() in [5, 6]:
                continue
            date_index = index + 4
            total_over_time += row[date_index].value
        # 小数
        xs = total_over_time - math.floor(total_over_time)
        # 根据小数对总加班时长四舍五入取整
        if xs < 0.5:
            total_over_time = math.floor(total_over_time)
        else:
            total_over_time = math.floor(total_over_time) + 1
        row[len(date_list) + 4].value = total_over_time
        # 统计周末加班时长
        for row in rows_name_list:
            total_over_time = 0
            for index in range(0, len(date_list)):
                day = date_list[index].date()
                if day.weekday() in [5, 6]:
                    date_index = index + 4
                    total_over_time += row[date_index].value
            # 小数
            xs = total_over_time - math.floor(total_over_time)
            # 根据小数对总加班时长四舍五入取整
            if xs < 0.5:
                total_over_time = math.floor(total_over_time)
            else:
                total_over_time = math.floor(total_over_time) + 1
            row[len(date_list) + 4 + 1].value = total_over_time

    # 生成最终结果时，'0'改为空值''
    column_num_summary = sheet_summary.max_column
    for row in list(sheet_summary.rows)[1:]:
        logger.info("row number: %s name: %s", row[2].value, row[3].value)
        for index in range(4, len(row)):
            logger.info("value: %s", row[index].value)
            value = row[index].value
            if value is not None and value == 0:
                row[index].value = ''

    file_calc_summary = os.path.join(os.getcwd(), FILE_DIR, FILE_SUMMARY)
    wb_summary.save(file_calc_summary)
    logger.info("第三次程序处理结果写入文件: %s", file_calc_summary)
