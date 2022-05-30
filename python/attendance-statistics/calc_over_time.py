#! /usr/bin/env python
# -*- coding: utf-8 -*-


import logging
import os
import openpyxl

from constant import *
from util_time import calc_duration_time

# pathname filename
logging.basicConfig(format='%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s',
                    level=logging.INFO)
logger = logging.getLogger(__name__)

if __name__ == '__main__':
    file_calc = os.path.join(os.getcwd(), FILE_DIR, FILE_PRE)
    logger.info("file path: %s", file_calc)
    wb = openpyxl.load_workbook(file_calc)
    sheet = wb.active
    row_num = sheet.max_row
    column_num = sheet.max_column
    logger.info("row number: %d, column number: %d", row_num, column_num)

    # 有效数据列需排除第一列
    rows_list = list(sheet.rows)[1:]

    # 计算周内加班时间
    for row in rows_list:
        if row[ORIGIN_COLUMN_IS_WEEKEND].value:
            # 跳过周末加班数据
            continue

        overtime = row[ORIGIN_COLUMN_OVER_TIME].value
        if overtime is None:
            # 不存在加班打卡记录
            row[ORIGIN_COLUMN_OVER_TIME_STATISTIC].value = 0
        else:
            overtime_list = [str_time.strip() for str_time in overtime.split(TIME_SEPARATOR)]
            overtime_list = [time_str for time_str in overtime_list if len(time_str) > 0]
            if len(overtime_list) != 2:
                logger.info("异常加班时间: %s", overtime_list)
                row[ORIGIN_COLUMN_COMMENT].value = "打卡次数异常(周内必须2次加班打卡记录)"
                row[ORIGIN_COLUMN_OVER_TIME_STATISTIC].value = -1
                continue
            overtime_start = overtime_list[0]
            overtime_last = overtime_list[1]
            if overtime_start < TIME_WEEKDAY_OVER_TIME_START:
                overtime_start = TIME_WEEKDAY_OVER_TIME_START

            overtime_statistic = calc_duration_time(overtime_start, overtime_last)
            row[ORIGIN_COLUMN_OVER_TIME_STATISTIC].value = overtime_statistic
            row[ORIGIN_COLUMN_OVER_TIME].value = TIME_SEPARATOR.join([overtime_start, overtime_last])

    # 计算周末加班时间
    for row in rows_list:
        if not row[ORIGIN_COLUMN_IS_WEEKEND].value:
            # 跳过周内加班数据
            continue

        overtime = row[ORIGIN_COLUMN_OVER_TIME].value
        if overtime is None:
            # 不存在加班打卡记录
            row[ORIGIN_COLUMN_OVER_TIME_STATISTIC].value = 0
        else:
            overtime_list = [str_time.strip() for str_time in overtime.split(TIME_SEPARATOR)]
            overtime_list = [time_str for time_str in overtime_list if len(time_str) > 0]
            overtime_statistic = 0
            if len(overtime_list) == 6:
                if overtime_list[2] <= TIME_WEEKEND_ON_WORK_AFTERNOON:
                    overtime_list[2] = TIME_WEEKEND_ON_WORK_AFTERNOON
                if overtime_list[4] <= TIME_WEEKDAY_OVER_TIME_NIGHT_START:
                    overtime_list[4] = TIME_WEEKDAY_OVER_TIME_NIGHT_START
                overtime_statistic = calc_duration_time(overtime_list[0], overtime_list[1]) \
                                     + calc_duration_time(overtime_list[2], overtime_list[3]) \
                                     + calc_duration_time(overtime_list[4], overtime_list[5])
            elif len(overtime_list) == 4:
                if overtime_list[2] <= TIME_WEEKEND_ON_WORK_AFTERNOON:
                    overtime_list[2] = TIME_WEEKEND_ON_WORK_AFTERNOON
                overtime_statistic = calc_duration_time(overtime_list[0], overtime_list[1]) \
                                     + calc_duration_time(overtime_list[2], overtime_list[3])
            elif len(overtime_list) == 2:
                overtime_statistic = calc_duration_time(overtime_list[0], overtime_list[1])
            else:
                logger.info("异常加班时间: %s", overtime_list)
                row[ORIGIN_COLUMN_COMMENT].value = "周末加班打卡次数异常(正常次数为2,4,6)"
                overtime_statistic = -1
            row[ORIGIN_COLUMN_OVER_TIME_STATISTIC].value = overtime_statistic
            row[ORIGIN_COLUMN_OVER_TIME].value = TIME_SEPARATOR.join(overtime_list)

    file_calc_summary = os.path.join(os.getcwd(), FILE_DIR, FILE_CALC)
    wb.save(file_calc_summary)
    logger.info("第二次程序处理结果写入文件: %s", file_calc_summary)
