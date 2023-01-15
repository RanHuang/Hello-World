#! /usr/bin/env python
# -*- coding: utf-8 -*-


import logging
import os
import openpyxl

from constant import *

# pathname filename
logging.basicConfig(format='%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s',
                    level=logging.INFO)
logger = logging.getLogger(__name__)


if __name__ == '__main__':
    file_excel_origin = os.path.join(os.getcwd(), FILE_DIR, FILE_ORIGIN_WEEKEND)
    logger.info("file path: %s", file_excel_origin)
    wb = openpyxl.load_workbook(file_excel_origin)
    sheet = wb.active
    row_num = sheet.max_row
    column_num = sheet.max_column
    logger.info("row number: %d, column number: %d", row_num, column_num)

    # 有效数据列需排除第一列
    rows_list = list(sheet.rows)[1:]

    logger.debug("print data of the first row:")
    for cell in rows_list[0]:
        logger.debug("type: %s, data: %s", type(cell.value), cell.value)

    # 提取加班打卡时间提取
    for row in rows_list:
        # 初始化备注列为空字符串
        row[ORIGIN_COLUMN_COMMENT].value = ""

        check_in_time = row[ORIGIN_COLUMN_CHECK_IN_TIME].value
        time_list = [str_time.strip() for str_time in check_in_time.split(TIME_SEPARATOR)]
        time_list = [time_str for time_str in time_list if len(time_str) > 0]
        logger.debug("check in time: %s", time_list)
        # 凌晨加班判定
        if time_list[0] < TIME_BEFORE_DOWN_END:
            row[ORIGIN_COLUMN_IS_BEFORE_DOWN].value = STRING_YES
            row[ORIGIN_COLUMN_COMMENT].value += "发现凌晨打卡；"

        if row[ORIGIN_COLUMN_IS_WEEKEND].value:
            # 周末加班
            overtime_list = time_list
            row[ORIGIN_COLUMN_OVER_TIME].value = "; ".join(overtime_list)
            if len(overtime_list) % 2 != 0:
                row[ORIGIN_COLUMN_COMMENT].value += "周末奇数次打卡；"
            if len(overtime_list) == 2:
                row[ORIGIN_COLUMN_COMMENT].value += "周日仅2次打卡；"
        else:
            # 过滤17:50之后的工作日打卡时间
            time_list_after_work = [time_str for time_str in time_list if time_str > TIME_WEEKDAY_OFF_TIME]
            # logger.info("weekday over time: %s", time_list_after_work)
            overtime_list = []
            if len(time_list_after_work) > 0:
                if time_list_after_work[-1] <= TIME_WEEKDAY_OVER_TIME_VALID:
                    # 打卡时间为19:00之前，算作无加班数据
                    overtime_list = []
                else:
                    # 加班时间修正为：18:30和最后一次打卡时间
                    overtime_list = [TIME_WEEKDAY_OVER_TIME_START, time_list_after_work[-1]]
            else:
                time_list_after_work = []
            logger.info("weekday over time origin vs fix: %s -- %s", time_list_after_work, overtime_list)
            row[ORIGIN_COLUMN_OVER_TIME].value = "; ".join(overtime_list)

    file_calc_origin = os.path.join(os.getcwd(), FILE_DIR, FILE_PRE)
    wb.save(file_calc_origin)
    logger.info("第一次程序处理结果写入文件: %s", file_calc_origin)
