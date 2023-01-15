#! /usr/bin/env python
# -*- coding: utf-8 -*-

import logging
import os
import sys
import openpyxl

from constant import *

# pathname filename
logging.basicConfig(format='%(asctime)s - %(filename)s[line:%(lineno)d] - %(levelname)s: %(message)s',
                    level=logging.INFO)
logger = logging.getLogger(__name__)

COLOR_RED = openpyxl.styles.colors.Color(rgb='00FF0000')
PATTERN_FILL = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=COLOR_RED)

if __name__ == '__main__':
    logger.info("根据%s的异常数据彩色标注%s，输出到新文件%s", FILE_CALC, FILE_PRE, FILE_PRE_COLOR)

    file_pre = os.path.join(os.getcwd(), FILE_DIR, FILE_PRE)
    logger.info("file path: %s", file_pre)
    wb_pre = openpyxl.load_workbook(file_pre)
    sheet_pre = wb_pre.active
    logger.info("%s row number: %d, column number: %d", FILE_PRE, sheet_pre.max_row, sheet_pre.max_column)

    file_calc = os.path.join(os.getcwd(), FILE_DIR, FILE_CALC)
    logger.info("file path: %s", file_calc)
    wb_calc = openpyxl.load_workbook(file_calc)
    sheet_calc = wb_calc.active
    logger.info("%s row number: %d, column number: %d", FILE_CALC, sheet_calc.max_row, sheet_calc.max_column)

    if sheet_calc.max_row != sheet_pre.max_row:
        logger.error("row number is not equal in %s and %s", FILE_CALC, FILE_PRE)
        sys.exit(1)
    if sheet_calc.max_column != sheet_pre.max_column:
        logger.error("column number is not equal in %s and %s", FILE_CALC, FILE_PRE)
        sys.exit(1)

    # 有效数据列需排除第一列
    rows_calc_list = list(sheet_calc.rows)[1:]
    rows_pre_list = list(sheet_pre.rows)[1:]
    # 并行迭代两个列表
    index = 1
    for row_calc, row_pre in zip(rows_calc_list, rows_pre_list):
        index = index + 1
        overtime_statistic = row_calc[ORIGIN_COLUMN_OVER_TIME_STATISTIC].value
        overtime = row_calc[ORIGIN_COLUMN_OVER_TIME].value
        if overtime_statistic < 0:
            logger.info("行%d, 异常加班时长: %s， 加班时间：%s。[加班时长 < 0]", index, overtime_statistic, overtime)
            row_pre[ORIGIN_COLUMN_CHECK_IN_TIME].fill = PATTERN_FILL
        if 0 < overtime_statistic < 0.5:
            logger.info("行%d, 异常加班时长: %s， 加班时间：%s。[加班时长 < 0.5]", index, overtime_statistic, overtime)
            row_pre[ORIGIN_COLUMN_CHECK_IN_TIME].fill = PATTERN_FILL
        if overtime_statistic > 12:
            logger.info("行%d, 异常加班时长: %s， 加班时间：%s。[加班时长 > 12]", index, overtime_statistic, overtime)
            row_pre[ORIGIN_COLUMN_CHECK_IN_TIME].fill = PATTERN_FILL
        if overtime_statistic == 0 and overtime is not None:
            logger.info("行%d，异常加班时长: %s， 加班时间：%s。[加班时长为0但是加班时间(H列)不是空值]", index, overtime_statistic, overtime)
            row_pre[ORIGIN_COLUMN_CHECK_IN_TIME].fill = PATTERN_FILL

    file_pre_color = os.path.join(os.getcwd(), FILE_DIR, FILE_PRE_COLOR)
    wb_pre.save(file_pre_color)
    logger.info("将%s中的异常数据彩色标注到%s", FILE_CALC, FILE_PRE_COLOR)
